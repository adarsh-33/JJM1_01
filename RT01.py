#!/usr/bin/env python
# coding: utf-8
"""
Started on Mon Sep  5 11:11:11 2022
Completed on Wed Sep 7 8:30:05 2022
@author: AdarshPradhan
"""
# In[1]:

import sys
import os 
import subprocess
from subprocess import STDOUT, check_call
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'xlsxwriter'])
#subprocess.check_call([sys.executable, '-y', 'sudo apt-get', 'install', 'python-tk'])
#check_call(['apt-get', 'install', '-y', 'python-tk'], stdout=open(os.devnull,'wb'), stderr=STDOUT) 
os.system('echo %s|sudo -S %s' % ('mypass', 'apt-get -y install tk'))
import time, sys
import pandas as pd
from IPython.display import display
#from ipyfilechooser import FileChooser
from IPython.display import clear_output
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
import xlsxwriter
import re
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename, askdirectory
import datetime

# In[2]:


# Retrieve current working directory (`cwd`)
cwd = os.getcwd()


# In[3]:


def update_progress(progress,n):
    bar_length = 7
    if isinstance(progress, int):
        progress = float(progress)
    if not isinstance(progress, float):
        progress = 0
    if progress < 0:
        progress = 0
    if progress >= 1:
        progress = 1

    block = int(round(bar_length * progress))

    clear_output(wait = True)
    text = "Reading "+n+" : [{0}] {1:.1f}%".format( "#" * block + "-" * (bar_length - block), progress * 100)
    print(text)


# In[4]:

Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
fc1 = askdirectory(title='Select the Directory for EAT02.xlsx Files',initialdir=cwd)
fc2 = askopenfilename(title='Select the EAT11.xlsx File', filetypes =[('Excel File', '*.xlsx')],initialdir=cwd)
fc3 = askdirectory(title='Select the Directory for EP04.xlsx Files',initialdir=cwd)
print("FC1",fc1)
print("FC2",fc2)
print("FC2",fc3)

# # Selecting the Directory fir EAT02 files
# # Create and display a FileChooser widget and Switch to folder-only mode
# fc1 = FileChooser(cwd)
# fc1.title = '<b>Select the Directory for EAT02.xlsx Files</b>'
# fc1.show_only_dirs = True
# display(fc1)


# In[5]:


# Create and display a FileChooser widget
# fc2 = FileChooser(cwd)
# fc2.title = '<b>Select the EAT11.xlsx File</b>'
# fc2.filter_pattern = '*.xlsx'

#fc1.reset(path="C:/Users/")
#fc2.reset(path="C:/Users/")
#fc3.reset(path="C:/Users/")

#display(fc2)


# In[6]:


# Selecting the Directory fir EAT02 files
# Create and display a FileChooser widget and Switch to folder-only mode
# print ("Select the Directory for EP04 xlsx Files")
# fc3 = FileChooser(cwd)
# fc3.title = '<b>Select the Directory for EP04.xlsx Files</b>'
# fc3.show_only_dirs = True
# display(fc3)


# In[7]:


## Reading all the EAT02.xlsx Files
cnt_eat02=0
sheet_eat02 = list()
#dir_path1 = os.listdir(fc1.selected)
#for x in os.listdir(fc1.selected):
for x in os.listdir(fc1):
    if x.endswith(".xlsx"):
        cnt_eat02+=1
        #print(fc1.selected+x)

        number_of_elements = 10
        for i in range(number_of_elements):
            # Load in the workbook
            update_progress(i / number_of_elements,x)
            wb1 = load_workbook(fc1+"/"+x)
            #wb1 = load_workbook(fc1.selected+x)

        update_progress(1,x)

        st1=wb1.sheetnames[0]
        sheet_eat02.append(wb1[st1])    

print(str(cnt_eat02)+" EAT02 Files Read.")


# In[8]:


## Reading EAT11.xlsx File
#file_path2=fc2.selected
file_path2=fc2
number_of_elements = 10
for i in range(number_of_elements):
    # Load in the workbook
    #update_progress(i / number_of_elements,fc2.selected_filename)
    update_progress(i / number_of_elements,fc2)
    wb2 = load_workbook(file_path2)

#update_progress(1,fc2.selected_filename)
update_progress(1,fc2)

st2=wb2.sheetnames[0]
sheet2 = wb2[st2]


# In[9]:


## Reading all the EP04.xlsx Files
cnt_ep04=0
sheet_ep04 = list()
#dir_path1 = os.listdir(fc3.selected)
#for x in os.listdir(fc3.selected):
for x in os.listdir(fc3):
    if x.endswith(".xlsx"):
        cnt_ep04+=1
        number_of_elements = 10
        for i in range(number_of_elements):
            # Load in the workbook
            update_progress(i / number_of_elements,x)
            #wb1 = load_workbook(fc3.selected+x)
            wb1 = load_workbook(fc3+"/"+x)

        update_progress(1,x)

        st1=wb1.sheetnames[0]
        sheet_ep04.append(wb1[st1])    

print(str(cnt_ep04)+" EP04 Files Read.")


# In[11]:


##Extracting columns from EAT11 xlsx

## 1)  'Office order no' ROW -> eat11_oford_row
##     'Office order no' COLUMN -> eat11_oford_col
## 2)  'Voucher number' ROW -> eat11_vochn_row
##     'Voucher number' COLUMN -> eat11_vochn_col
## 3)  'Voucher date' ROW -> eat11_vochdt_row
##     'Voucher date' COLUMN -> eat11_vochdt_col
## 4)  'Voucher amount (Total Amount of Voucher)' ROW -> eat11_vochamt_row
##     'Voucher amount (Total Amount of Voucher)' COLUMN -> eat11_vochamt_col
## 5)  'PPA/Cheque number/ PFMS Batch no. (in case of DSC)' ROW -> eat11_chqno_row
##     'PPA/Cheque number/ PFMS Batch no. (in case of DSC)' COLUMN -> eat11_chqno_col
## 6)  'PPA/Cheque amount/ PFMS Batch amount (in case of DSC)' ROW -> eat11_chqamt_row
##     'PPA/Cheque amount/ PFMS Batch amount (in case of DSC)' COLUMN -> eat11_chqamt_col
## 7)  'Deduction amount' ROW -> eat11_deamt_row
##     'Deduction amount' COLUMN -> eat11_deamt_col
## 8)  'Deduction type' ROW -> eat11_detyp_row
##     'Deduction type' COLUMN ->eat11_detyp _col

lst_row_eat11=sheet2.max_row
lst_col_eat11=sheet2.max_column
##1) Search for Column with 'Office order no'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="Office order no.".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                #print(sheet2.cell(row=i, column=j).value)
                #print(i, j)
                eat11_oford_row=i
                eat11_oford_col=j
                extt=1
                break
    if extt==1:
        break
##2) Search for Column with 'Voucher number'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="Voucher number".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_vochn_row=i
                eat11_vochn_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_vochn_row, eat11_vochn_col)
                extt=1
                break
    if extt==1:
        break
##3) Search for Column with 'Voucher date'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="Voucher date".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_vochdt_row=i
                eat11_vochdt_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_vochdt_row, eat11_vochdt_col)
                extt=1
                break
    if extt==1:
        break
##4) Search for Column with 'Voucher amount (Total Amount of Voucher)'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="Voucher amount (Total Amount of Voucher)".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_vochamt_row=i
                eat11_vochamt_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_vochamt_row, eat11_vochamt_col)
                extt=1
                break
    if extt==1:
        break
##5) Search for Column with 'PPA/Cheque number/ PFMS Batch no. (in case of DSC)'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="PPA/Cheque number/ PFMS Batch no. (in case of DSC)".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_chqno_row=i
                eat11_chqno_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_chqno_row, eat11_chqno_col)
                extt=1
                break
    if extt==1:
        break
##6) Search for Column with 'PPA/Cheque amount/ PFMS Batch amount (in case of DSC)'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="PPA/Cheque amount/ PFMS Batch amount (in case of DSC)".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_chqamt_row=i
                eat11_chqamt_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_chqamt_row, eat11_chqamt_col)
                #print(str1, str2)
                extt=1
                break
    if extt==1:
        break
##7) Search for Column with 'Deduction amount'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="Deduction amount".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_deamt_row=i
                eat11_deamt_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_deamt_row, eat11_deamt_col)
                #print(str1, str2)
                extt=1
                break
    if extt==1:
        break
##8) Search for Column with 'Deduction type'
extt=0
for i in range(1,lst_row_eat11):
    for j in range (1,lst_col_eat11):
        if (sheet2.cell(row=i, column=j).value!=None):
            str1=str((sheet2.cell(row=i, column=j).value)).lower()
            str2="Deduction type".lower()
            str1=re.sub('\s+', '', str1.strip())
            str2=re.sub('\s+', '', str2.strip())
            if (str1==str2):
                eat11_detyp_row=i
                eat11_detyp_col=j
                #print(sheet2.cell(row=i, column=j).value)
                #print(eat11_detyp_row, eat11_detyp_col)
                #print(str1, str2)
                extt=1
                break
    if extt==1:
        break

# print ('MAX ROW in EAT11:- ',lst_row_eat11,'MAX COLUMN in EAT11:- ',lst_col_eat11)
# print ('\nOffice order no', eat11_oford_row, eat11_oford_col,'\nVoucher number',eat11_vochn_row
#       ,eat11_vochn_col,'\nVoucher date',eat11_vochdt_row,eat11_vochdt_col,
#       '\nVoucher amount (Total Amount of Voucher)',eat11_vochamt_row,eat11_vochamt_col,
#       '\nPPA/Cheque number/ PFMS Batch no. (in case of DSC)',eat11_chqno_row,eat11_chqno_col,
#       '\nPPA/Cheque amount/ PFMS Batch amount (in case of DSC)',eat11_chqamt_row,eat11_chqamt_col,
#       '\nDeduction amount',eat11_deamt_row,eat11_deamt_col,'\nDeduction type',eat11_detyp_row,eat11_detyp_col)


# In[12]:


##Extracting columns from EAT02 xlsx 
eat02_sanct=list()
eat02_vochn=list()
eat02_recpt=list()
for ind in range (0,len(sheet_eat02)):
    ##Search for Column with 'Sanction No'
    extt=0
    for i in range(1,sheet_eat02[ind].max_row):
        for j in range (1,sheet_eat02[ind].max_column):
            if (sheet_eat02[ind].cell(row=i, column=j).value!=None):
                str1=str((sheet_eat02[ind].cell(row=i, column=j).value)).lower()
                str2="Sanction No".lower()
                str1=re.sub('\s+', '', str1.strip())
                str2=re.sub('\s+', '', str2.strip())
                if (str1==str2):
                    #print(sheet_eat02[ind].cell(row=i, column=j).value)
                    #print(i, j)
                    eat02_sanct.append([i,j])
#                     eat02_sanct_row=i
#                     eat02_sanct_col=j
                    extt=1
                    break
        if extt==1:
            break
                
    ##Search for Column with 'Voucher number'
    extt=0
    for i in range(1,sheet_eat02[ind].max_row):
        for j in range (1,sheet_eat02[ind].max_column):
            if (sheet_eat02[ind].cell(row=i, column=j).value!=None):
                str1=str((sheet_eat02[ind].cell(row=i, column=j).value)).lower()
                str2="Voucher number".lower()
                str1=re.sub('\s+', '', str1.strip())
                str2=re.sub('\s+', '', str2.strip())
                if (str1==str2):
                    #print(sheet_eat02[ind].cell(row=i, column=j).value)
                    #print(i, j)
                    eat02_vochn.append([i,j])
#                     eat02_vochn_row=i
#                     eat02_vochn_col=j
                    extt=1
                    break
        if extt==1:
            break

    ##Search for Column with 'Recipient'
    extt=0
    for i in range(1,sheet_eat02[ind].max_row):
        for j in range (1,sheet_eat02[ind].max_column):
            if (sheet_eat02[ind].cell(row=i, column=j).value!=None):
                str1=str((sheet_eat02[ind].cell(row=i, column=j).value)).lower()
                str2="Recipient".lower()
                str1=re.sub('\s+', '', str1.strip())
                str2=re.sub('\s+', '', str2.strip())
                if (str1==str2):
                    #print(sheet_eat02[ind].cell(row=i, column=j).value)
                    #print(i, j)
#                     eat02_recpt_row=i
#                     eat02_recpt_col=j
                    eat02_recpt.append([i,j])
                    extt=1
                    break
        if extt==1:
            break
#print("Sanction No",eat02_sanct,"\nVoucher number",eat02_vochn,"\nRecipient",eat02_recpt)


# In[13]:


##Extracting columns from EP04 xlsx
ep04_agncy=list()
ep04_debtn=list()
ep04_vochn=list()
for ind in range (0,len(sheet_ep04)):
    ##Search for Column with 'Debit Agency Name'
    extt=0
    for i in range(1,sheet_ep04[ind].max_row):
        for j in range (1,sheet_ep04[ind].max_column):
            if (sheet_ep04[ind].cell(row=i, column=j).value!=None):
                str1=str((sheet_ep04[ind].cell(row=i, column=j).value)).lower()
                str2="Debit Agency Name".lower()
                str1=re.sub('\s+', '', str1.strip())
                str2=re.sub('\s+', '', str2.strip())
                if (str1==str2):
                    #print(sheet_ep04[ind].cell(row=i, column=j).value)
                    #print(i, j)
                    ep04_agncy.append([i,j])
#                     ep04_agncy_row=i
#                     ep04_agncy_col=j
                    extt=1
                    break
        if extt==1:
            break
    ##Search for Column with 'Debit Batch No/Advice No.'
    extt=0
    for i in range(1,sheet_ep04[ind].max_row):
        for j in range (1,sheet_ep04[ind].max_column):
            if (sheet_ep04[ind].cell(row=i, column=j).value!=None):
                str1=str((sheet_ep04[ind].cell(row=i, column=j).value)).lower()
                str2="Debit Batch No/Advice No.".lower()
                str1=re.sub('\s+', '', str1.strip())
                str2=re.sub('\s+', '', str2.strip())
                if (str1==str2):
                    #print(sheet_ep04[ind].cell(row=i, column=j).value)
                    #print(i, j)
                    ep04_debtn.append([i,j])
#                     ep04_debtn_row=i
#                     ep04_debtn_col=j
                    extt=1
                    break
        if extt==1:
            break
    ##Search for Column with 'Debit Voucher No.'
    extt=0
    for i in range(1,sheet_ep04[ind].max_row):
        for j in range (1,sheet_ep04[ind].max_column):
            if (sheet_ep04[ind].cell(row=i, column=j).value!=None):
                str1=str((sheet_ep04[ind].cell(row=i, column=j).value)).lower()
                str2="Debit Voucher No.".lower()
                str1=re.sub('\s+', '', str1.strip())
                str2=re.sub('\s+', '', str2.strip())
                if (str1==str2):
                    #print(sheet_ep04[ind].cell(row=i, column=j).value)
                    #print(i, j)
                    ep04_vochn.append([i,j])
#                     ep04_vochn_row=i
#                     ep04_vochn_col=j
                    extt=1
                    break
        if extt==1:
            break
#print("Debit Agency Name ",ep04_agncy,"\nDebit Batch No/Advice No ",ep04_debtn,"\nDebit Voucher No ",ep04_vochn)


# In[93]:


##Retrieving data from EAT11 xlsx File
## 1)  'Office order no' ROW -> eat11_oford_row
##     'Office order no' COLUMN -> eat11_oford_col
## 2)  'Voucher number' ROW -> eat11_vochn_row
##     'Voucher number' COLUMN -> eat11_vochn_col
## 3)  'Voucher date' ROW -> eat11_vochdt_row
##     'Voucher date' COLUMN -> eat11_vochdt_col
## 4)  'Voucher amount (Total Amount of Voucher)' ROW -> eat11_vochamt_row
##     'Voucher amount (Total Amount of Voucher)' COLUMN -> eat11_vochamt_col
## 5)  'PPA/Cheque number/ PFMS Batch no. (in case of DSC)' ROW -> eat11_chqno_row
##     'PPA/Cheque number/ PFMS Batch no. (in case of DSC)' COLUMN -> eat11_chqno_col
## 6)  'PPA/Cheque amount/ PFMS Batch amount (in case of DSC)' ROW -> eat11_chqamt_row
##     'PPA/Cheque amount/ PFMS Batch amount (in case of DSC)' COLUMN -> eat11_chqamt_col
## 7)  'Deduction amount' ROW -> eat11_deamt_row
##     'Deduction amount' COLUMN -> eat11_deamt_col
## 8)  'Deduction type' ROW -> eat11_detyp_row
##     'Deduction type' COLUMN ->eat11_detyp_col

val_eat11 = list()
strt_row_eat11=eat11_oford_row+3
a1="Voucher number"
a2="Voucher date"
a3="Voucher amount (Gross Amount)"
a4="Office order no.(Sanction No)"
a5="PPA/Cheque number/ PFMS Batch no."
a6="PPA/Cheque amount/ PFMS Batch amount (Net Amount)"
a7="Deduction Amount"#AMT, TYPE -- AMT, TYPE 
a8="Deduction Type"#AMT, TYPE -- AMT, TYPE 
a9="Total Deduction Amount"
a10="Nos. of Deductions"

val_eat11.append([a1,a2,a3,a4,a5,a6,a7,a8,a9,a10])

for ind in range(strt_row_eat11,lst_row_eat11) :
    if(sheet2.cell(row=ind, column=eat11_vochn_col).value!=None):
        val2 = list()
        a1=sheet2.cell(row=ind, column=eat11_vochn_col).value
        val2.append(a1)
        a1=sheet2.cell(row=ind, column=eat11_vochdt_col).value
        val2.append(a1)
        a1=float(sheet2.cell(row=ind, column=eat11_vochamt_col).value)
        val2.append(a1)
        a1=sheet2.cell(row=ind, column=eat11_oford_col).value
        val2.append(a1)
        a1=sheet2.cell(row=ind, column=eat11_chqno_col).value
        val2.append(a1)
        a1=float(sheet2.cell(row=ind, column=eat11_chqamt_col).value)
        val2.append(a1)
        #SUMMING THE DEDUC AMT and TAKING THE DEDUC AMT & TYPE
        ddam=0
        tr=ind
        cnt_dtp=0
        while ((sheet2.cell(row=tr+1, column=eat11_vochn_col).value==None) and (tr<lst_row_eat11-2)):
            ddam=ddam+float(sheet2.cell(row=tr, column=eat11_deamt_col).value)
            a1=float(sheet2.cell(row=tr, column=eat11_deamt_col).value)
            val2.append(a1)
            a1=sheet2.cell(row=tr, column=eat11_detyp_col).value
            val2.append(a1)
            tr+=1
            cnt_dtp+=1
        a1=float(sheet2.cell(row=tr, column=eat11_deamt_col).value)
        val2.append(a1)
        a1=sheet2.cell(row=tr, column=eat11_detyp_col).value
        val2.append(a1)
        ddam=ddam+float(sheet2.cell(row=tr, column=eat11_deamt_col).value)
        # dtp=dtp+sheet2.cell(row=tr, column=eat11_detyp_col).value
        val2.append(ddam)
        val2.append(cnt_dtp+1)
        
        val_eat11.append(val2)

val_eat11[0].insert(0,"Agency Name")#Inserting in 0th Row, 0th column
val_eat11[0].insert(3,"Recipient")#Inserting in 0th Row, 3rd column
#display(val_eat11)


# In[22]:


##Retrieving data from EP04 xlsx File
# ep04_vochn
# ep04_debtn
# ep04_agncy
val_ep04=list()
for stobj in range (0,len(sheet_ep04)):
    strt_row=ep04_vochn[stobj][0]+2 #13+2
    lst_row=sheet_ep04[stobj].max_row
    for ind in range(strt_row,lst_row) :
        if(sheet_ep04[stobj].cell(row=ind, column=ep04_vochn[stobj][1]).value!=None):
            val2 = list()
            a1=sheet_ep04[stobj].cell(row=ind, column=ep04_vochn[stobj][1]).value
            val2.append(a1)
            a1=sheet_ep04[stobj].cell(row=ind, column=ep04_debtn[stobj][1]).value
            val2.append(a1)
            a1=sheet_ep04[stobj].cell(row=ind, column=ep04_agncy[stobj][1]).value
            val2.append(a1)
            
            val_ep04.append(val2)
# display(val_ep04)


# In[23]:


##Retrieving data from EAT02 xlsx File
# eat02_vochn
# eat02_sanct
# eat02_recpt
val_eat02=list()
for stobj in range (0,len(sheet_eat02)):
    strt_row=eat02_sanct[stobj][0]+2 #13+2
    lst_row=sheet_eat02[stobj].max_row
    for ind in range(strt_row,lst_row) :
        if(sheet_eat02[stobj].cell(row=ind, column=eat02_sanct[stobj][1]).value!=None):
            val2 = list()
            a1=sheet_eat02[stobj].cell(row=ind, column=eat02_vochn[stobj][1]).value
            val2.append(a1)
            a1=sheet_eat02[stobj].cell(row=ind, column=eat02_sanct[stobj][1]).value
            val2.append(a1)
            a1=sheet_eat02[stobj].cell(row=ind, column=eat02_recpt[stobj][1]).value
            val2.append(a1)
            
            val_eat02.append(val2)
# display(val_eat02)


# In[25]:


# EAT11(val_eat11->0th column): a1="Voucher number" -->> In  val_ep04, 1st column, In val_eat02, 0th column
# EAT11(val_eat11->3rd column): a4="Office order no.(Sanction No)" -->> In val_eat02, 1st column
# EAT11(val_eat11->4th column): a5="PPA/Cheque number/ PFMS Batch no."  -->> In  val_ep04, 1st column
# EAT02: val_eat02 -->> [0,1,2]2nd column is the Recipient
# EP04: val_ep04 -->> [0,1,2]2nd column is the Agency Name
##How to insert an element in a list arr[1].insert(2,12)-> inserting '12' in 2nd row, 3rd column

for stobj in range (1 , len(val_eat11)):
    fnd=0
    for ind in range (0 , len(val_ep04)):
#         str1=str(val_eat11[stobj][0]).lower()
#         str2=str(val_eat11[stobj][4]).lower()
#         str3=str(val_ep04[ind][0]).lower()
#         str4=str(val_ep04[ind][1]).lower()
#         str1=re.sub('\s+', '', str1.strip())
#         str2=re.sub('\s+', '', str2.strip())
#         str3=re.sub('\s+', '', str3.strip())
#         str4=re.sub('\s+', '', str4.strip())
        
        #print("str1->",str1,"str3->", str3, "str2->",str2, "str4->",str4)
        if (val_eat11[stobj][0]==val_ep04[ind][0] and val_eat11[stobj][4]==val_ep04[ind][1]):
        #if (str1==str3 and str2==str4):
            fnd+=1
            val_eat11[stobj].insert(0,val_ep04[ind][2])
    if (fnd==0):
        val_eat11[stobj].insert(0,"")

for stobj in range (1 , len(val_eat11)):
    fnd=0
    for ind in range (0 , len(val_eat02)):
#         str1=str(val_eat11[stobj][1]).lower()
#         str2=str(val_eat11[stobj][4]).lower()
#         str3=str(val_eat02[ind][0]).lower()
#         str4=str(val_eat02[ind][1]).lower()
#         str1=re.sub('\s+', '', str1.strip())
#         str2=re.sub('\s+', '', str2.strip())
#         str3=re.sub('\s+', '', str3.strip())
#         str4=re.sub('\s+', '', str4.strip())

        #print("str1->",str1,"str3->", str3, "str2->",str2, "str4->",str4)
        if (val_eat11[stobj][1]==val_eat02[ind][0] and val_eat11[stobj][4]==val_eat02[ind][1]):
        #if (str1==str3 and str2==str4):
            #print(val_eat02[ind][2])
            fnd+=1
            val_eat11[stobj].insert(3,val_eat02[ind][2])
    if (fnd==0):
        val_eat11[stobj].insert(3,"")


# In[91]:


# Col 0 = "Agency Name"
# Col 1="Voucher number"
# Col 2="Voucher date"
# Col 3="Recipient"
# Col 4="Voucher amount (Gross Amount)"
# Col 5="Office order no.(Sanction No)"
# Col 6="PPA/Cheque number/ PFMS Batch no."
# Col 7="PPA/Cheque amount/ PFMS Batch amount (Net Amount)"
# Col 8 (HERE MERGING)="Deduction Amount"#AMT, TYPE -- AMT, TYPE 
# Col 9 (HERE MERGING)="Deduction Type"#AMT, TYPE -- AMT, TYPE 
# Col 10="Total Deduction Amount"
# Not Required -> Col 12="Nos. of Deductions"


time = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
name = 'JJM_Report %s.xlsx' % (time)
with xlsxwriter.Workbook(name) as workbook:
    worksheet = workbook.add_worksheet()
    #cell_format0 = workbook.add_format({'bold': True, 'font_size': 11,'align': 'center'})
    cell_format0 = workbook.add_format({'bold': True, 'font_size': 11, 'text_wrap': True, 'border': 2})
    cell_format1 = workbook.add_format({'align': 'top', 'border': 2})
    number_format = workbook.add_format({'num_format': '#,##0.00','align': 'top', 'border': 2})
    worksheet.freeze_panes(1, 0)
    worksheet.set_row(0, None, cell_format0)
    worksheet.set_column(0, 0, 40,cell_format1)
    worksheet.set_column(1, 1, 14,cell_format1)
    worksheet.set_column(2, 2, 17,cell_format1)
    worksheet.set_column(3, 3, 21,cell_format1)
    worksheet.set_column(4, 4, 15, number_format)
    worksheet.set_column(5, 5, 21,cell_format1)
    worksheet.set_column(6, 6, 17,cell_format1)
    worksheet.set_column(7, 7, 15, number_format)
    worksheet.set_column(8, 8, 11, number_format)
    worksheet.set_column(9, 9, 11,cell_format1)
    worksheet.set_column(10, 10, 11, number_format)
    pre_r=0
    for row_num, data in enumerate(val_eat11):
        if (row_num==0):
            for i in range (len(data)-1): #for data[]
                worksheet.write(row_num, i, data[i])
            pre_r+=1
        else:
            no_ele=len(data)
            no_deduc=data[len(data)-1]
            for i in range (8): #for data[]
                if (no_deduc==1):
                    worksheet.write(pre_r, i, data[i])
                else:
                    worksheet.merge_range(pre_r, i, pre_r+no_deduc-1, i, data[i])
            i=8
            for j in range (pre_r, pre_r+no_deduc):
                worksheet.write(j, 8, data[i])
                i+=1
                worksheet.write(j, 9, data[i])
                i+=1
            if (no_deduc==1):
                worksheet.write(pre_r, 10, data[i])
            else:
                worksheet.merge_range(pre_r, 10, pre_r+no_deduc-1, 10, data[i])
            pre_r+=no_deduc
            
#         else:
#             for i in range (len(data)-1): #for data[]
#                 if (i<=7):
#                     worksheet.merge_range(pre_r, i, data[len(data)-1]-1, i, data[i])
#                 else:
#                     worksheet.write(pre_r, 8, data[i])
#                     i+=1
#                     worksheet.write(pre_r, 9, data[i])
#                     pre_r+=1
                    #worksheet.write_row(row_num, 0, data)

print("FILE DOWNLOADED SUCCESSFULLY!!")

